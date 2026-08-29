from datetime import date
from decimal import Decimal, InvalidOperation
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from backend.models import Category, Department, IssuanceLine, IssuanceRecord, StockItem


BANGLA_DIGITS = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
MONTHS = {
    "জুলাই": "July",
    "আগস্ট": "August",
    "সেপ্টেম্বর": "September",
    "অক্টোবর": "October",
    "নভেম্বর": "November",
    "ডিসেম্বর": "December",
    "জানুয়ারি": "January",
    "জানুয়ারি": "January",
    "ফেব্রুয়ারি": "February",
    "ফেব্রুয়ারি": "February",
    "মার্চ": "March",
    "এপ্রিল": "April",
    "মে": "May",
    "জুন": "June",
}
UNIT_MAP = {
    "রিম": "ream",
    "টি": "piece",
    "কেজি": "kg",
    "সেট": "set",
    "বক্স": "box",
    "প্যাকেট": "packet",
    "রোল": "roll",
}


class Command(BaseCommand):
    help = "Import stock issuance records from the final store calculation workbook."

    panel_ranges = (
        (1, 2, 3, 43),     # A:AQ, items 1-40, department AQ
        (51, 52, 53, 90),  # AY:CL, items 41-77, department CL
        (98, 99, 100, 141),  # CT:EK, items 78-118, department EK
    )

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to the .xlsx workbook")
        parser.add_argument("--sheet", default=None, help="Worksheet name. Defaults to the active sheet.")
        parser.add_argument("--fiscal-year", default="2025-26")
        parser.add_argument("--dry-run", action="store_true", help="Parse and report without saving data.")
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Delete existing issuance records for the fiscal year before importing.",
        )

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        fiscal_year = options["fiscal_year"]

        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"Workbook not found: {workbook_path}") from exc

        worksheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        sections = self.find_sections(rows)

        if not sections:
            raise CommandError("No month/sheet sections were found in the workbook.")

        stats = {"categories": 0, "items": 0, "departments": 0, "records": 0, "lines": 0}

        context = transaction.atomic() if not options["dry_run"] else nullcontext()
        with context:
            if options["replace"] and not options["dry_run"]:
                IssuanceRecord.objects.filter(fiscal_year=fiscal_year).delete()

            for index, start_row in enumerate(sections):
                end_row = sections[index + 1] if index + 1 < len(sections) else len(rows)
                section_stats = self.import_section(rows, start_row, end_row, fiscal_year, options["dry_run"])
                for key, value in section_stats.items():
                    stats[key] += value

        self.stdout.write(
            self.style.SUCCESS(
                "Imported workbook: "
                f"{stats['categories']} categories, {stats['items']} items, "
                f"{stats['departments']} departments, {stats['records']} records, "
                f"{stats['lines']} lines."
            )
        )

    def find_sections(self, rows):
        sections = []
        for index, row in enumerate(rows):
            value = self.clean_text(row[0] if row else "")
            if "শীট" in value and not value.startswith("শিক্ষা"):
                sections.append(index)
        return sections

    def import_section(self, rows, start_row, end_row, fiscal_year, dry_run):
        title = self.clean_text(rows[start_row][0])
        month = self.month_from_title(title)
        sheet_number = self.sheet_number_from_title(title)
        category_row = rows[start_row + 1] if start_row + 1 < len(rows) else ()
        item_row = rows[start_row + 2] if start_row + 2 < len(rows) else ()
        number_row = rows[start_row + 3] if start_row + 3 < len(rows) else ()
        unit_row = rows[start_row + 4] if start_row + 4 < len(rows) else ()

        stats = {"categories": 0, "items": 0, "departments": 0, "records": 0, "lines": 0}
        item_columns = {}
        dry_run_categories = set()

        for _serial_col, _date_col, first_item_col, department_col in self.panel_ranges:
            current_category = "Uncategorized"
            for column in range(first_item_col, department_col):
                category_name = self.clean_text(self.cell(category_row, column)) or current_category
                current_category = category_name
                item_name = self.clean_text(self.cell(item_row, column))
                item_number = self.to_int(self.cell(number_row, column))
                if not item_name or item_number is None:
                    continue

                unit = UNIT_MAP.get(self.clean_text(self.cell(unit_row, column)), "other")
                if not dry_run:
                    category, created = Category.objects.get_or_create(name=category_name)
                    stats["categories"] += int(created)
                    item, created = StockItem.objects.update_or_create(
                        item_number=item_number,
                        defaults={"name": item_name, "unit": unit, "category": category},
                    )
                    stats["items"] += int(created)
                    item_columns[column] = item
                else:
                    dry_run_categories.add(category_name)
                    stats["items"] += 1
                    item_columns[column] = item_number
        stats["categories"] += len(dry_run_categories)

        for row_index in range(start_row + 5, end_row):
            row = rows[row_index]
            records = {}
            for serial_col, date_col, first_item_col, department_col in self.panel_ranges:
                serial_number = self.to_int(self.cell(row, serial_col))
                record_date = self.parse_date(self.cell(row, date_col))
                department_name = self.clean_text(self.cell(row, department_col))
                if serial_number is None or record_date is None or not department_name:
                    continue

                key = (serial_number, record_date, department_name)
                quantities = records.setdefault(key, {})
                for column in range(first_item_col, department_col):
                    item = item_columns.get(column)
                    quantity = self.to_decimal(self.cell(row, column))
                    if item is not None and quantity is not None and quantity != 0:
                        quantities[item] = quantities.get(item, Decimal("0")) + quantity

            for (serial_number, record_date, department_name), quantities in records.items():
                if not quantities:
                    continue

                if dry_run:
                    stats["records"] += 1
                    stats["lines"] += len(quantities)
                    continue

                department, created = Department.objects.get_or_create(name=department_name)
                stats["departments"] += int(created)
                record, created = IssuanceRecord.objects.update_or_create(
                    serial_number=serial_number,
                    date=record_date,
                    department=department,
                    fiscal_year=fiscal_year,
                    sheet_number=sheet_number,
                    defaults={"month": month, "notes": title},
                )
                stats["records"] += 1

                record.lines.all().delete()
                IssuanceLine.objects.bulk_create(
                    [
                        IssuanceLine(record=record, item=item, quantity=quantity)
                        for item, quantity in quantities.items()
                    ]
                )
                stats["lines"] += len(quantities)

        return stats

    def cell(self, row, one_based_column):
        index = one_based_column - 1
        if index < 0 or index >= len(row):
            return None
        return row[index]

    def clean_text(self, value):
        if value is None:
            return ""
        return str(value).strip().translate(BANGLA_DIGITS)

    def to_int(self, value):
        text = self.clean_text(value)
        if not text:
            return None
        try:
            return int(Decimal(text))
        except (InvalidOperation, ValueError):
            return None

    def to_decimal(self, value):
        text = self.clean_text(value)
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def parse_date(self, value):
        if isinstance(value, date):
            return value
        text = self.clean_text(value)
        if not text:
            return None

        match = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", text)
        if not match:
            return None

        day, month, year = (int(part) for part in match.groups())
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None

    def month_from_title(self, title):
        for bangla, english in MONTHS.items():
            if bangla in title:
                return english
        return ""

    def sheet_number_from_title(self, title):
        text = title.translate(BANGLA_DIGITS)
        match = re.search(r"শীট\D*(\d+)", text)
        return int(match.group(1)) if match else None


class nullcontext:
    def __enter__(self):
        return None

    def __exit__(self, *exc_info):
        return False
